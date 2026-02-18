import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Параметры
NUM_CLIENTS = 1000
YEARS = [2023, 2024, 2025]
MONTHS = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
          'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь']
GROUPS = [f'Группа {i}' for i in range(1, 29)]  # 28 групп
METRICS = ['Количество в чеке', 'Сумма в чеке', 'Число чеков']

# Создаём workbook
wb = Workbook()
ws = wb.active
ws.title = 'Продажи'

# Вычисляем общее количество столбцов данных
num_metrics = len(METRICS)
num_groups = len(GROUPS)
num_months = len(MONTHS)
num_years = len(YEARS)

# Столбцов данных: годы * месяцы * группы * показатели
total_data_cols = num_years * num_months * num_groups * num_metrics

# Строка 1: Годы
col = 2  # начинаем со второго столбца (первый - ID клиента)
for year in YEARS:
    ws.cell(row=1, column=col, value=year)
    # Объединяем ячейки для года
    year_span = num_months * num_groups * num_metrics
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + year_span - 1)
    col += year_span

# Строка 2: Месяцы
col = 2
for year in YEARS:
    for month in MONTHS:
        ws.cell(row=2, column=col, value=month)
        # Объединяем ячейки для месяца
        month_span = num_groups * num_metrics
        ws.merge_cells(start_row=2, start_column=col, end_row=2, end_column=col + month_span - 1)
        col += month_span

# Строка 3: Группы товаров
col = 2
for year in YEARS:
    for month in MONTHS:
        for group in GROUPS:
            ws.cell(row=3, column=col, value=group)
            # Объединяем ячейки для группы
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + num_metrics - 1)
            col += num_metrics

# Строка 4: Показатели + заголовок ID клиента
ws.cell(row=4, column=1, value='ИД клиента')
col = 2
for year in YEARS:
    for month in MONTHS:
        for group in GROUPS:
            for metric in METRICS:
                ws.cell(row=4, column=col, value=metric)
                col += 1

# Генерация данных клиентов
np.random.seed(42)

print('Генерация данных...')
for client_idx in range(NUM_CLIENTS):
    client_id = f'CLIENT_{client_idx + 1:04d}'
    row = 5 + client_idx
    ws.cell(row=row, column=1, value=client_id)

    col = 2
    for year in YEARS:
        for month_idx, month in enumerate(MONTHS):
            for group_idx, group in enumerate(GROUPS):
                # Генерируем случайные данные с некоторой логикой
                # Некоторые клиенты не покупают некоторые группы (пропуски)
                if np.random.random() < 0.3:  # 30% пропусков
                    quantity = None
                    total = None
                    checks = None
                else:
                    quantity = np.random.randint(1, 50)
                    price_per_unit = np.random.uniform(100, 5000)
                    total = round(quantity * price_per_unit, 2)
                    checks = np.random.randint(1, min(quantity + 1, 10))

                ws.cell(row=row, column=col, value=quantity)
                ws.cell(row=row, column=col + 1, value=total)
                ws.cell(row=row, column=col + 2, value=checks)
                col += 3

    if (client_idx + 1) % 100 == 0:
        print(f'Обработано {client_idx + 1} клиентов из {NUM_CLIENTS}')

# Сохраняем файл
output_file = 'test_sales_data.xlsx'
print(f'Сохранение файла {output_file}...')
wb.save(output_file)
print(f'Файл {output_file} успешно создан!')
print(f'Размер: {NUM_CLIENTS} клиентов, {total_data_cols} столбцов данных')
